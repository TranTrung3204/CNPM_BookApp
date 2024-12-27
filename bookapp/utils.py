import openpyxl
from openpyxl.utils import get_column_letter
from sqlalchemy import func
from bookapp.models import BookCategory, Book, User, Receipt, ReceiptDetail, UserRole, DeliveryMethod, PaymentMethod
from bookapp import app, db
from flask_login import  current_user
import hashlib
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill



def load_book_categories():
    categories = BookCategory.query.order_by('id').all()
    if not categories:
        return []  # Đảm bảo trả về danh sách rỗng nếu không có danh mục
    for category in categories:
        category.product_count = Book.query.filter(Book.category_id == category.id).count()
    return categories



def load_books(kw: object = None) -> object:
    products = Book.query
    if kw:
        products = products.filter(Book.name.contains(kw))
    return products.all()


def add_user(name, username, password, **kwargs):
    password = str(hashlib.md5(password.strip().encode('utf-8')).hexdigest())
    user = User(name=name.strip(),
                username=username.strip(),
                password=password,
                email=kwargs.get('email'),
                avatar=kwargs.get('avatar'))
    db.session.add(user)
    db.session.commit()

def search_books(kw):
    if not kw:
        return []
    # Bỏ khoảng trắng thừa, chuyển về chữ thường, tìm kiếm
    return Book.query.filter(
        Book.name.ilike(f"%{kw}%")
    ).all()

def load_books_by_category(category_id):
    return Book.query.filter(Book.category_id == category_id).all()

def load_book_categories():
    categories = BookCategory.query.order_by('id').all()
    for category in categories:
        category.product_count = Book.query.filter(Book.category_id == category.id).count()
    return categories


def check_login(username, password, user_role=None):
    if username and password:
        password = str(hashlib.md5(password.strip().encode('utf-8')).hexdigest())
        query = User.query.filter(User.username.__eq__(username.strip()),
                                  User.password.__eq__(password))

        if user_role:
            if isinstance(user_role, list):
                # Nếu user_role là list thì kiểm tra user có thuộc một trong các role đó không
                query = query.filter(User.user_role.in_(user_role))
            else:
                # Nếu user_role là giá trị đơn lẻ
                query = query.filter(User.user_role.__eq__(user_role))

        return query.first()


def get_user_by_id(user_id):
    return User.query.get(user_id)

def count_cart(cart):
    total_quantity, total_amount = 0, 0
    if cart:
        for c in cart.values():
            c['price'] = float(c['price'])  # Chuyển price thành float
            c['quantity'] = int(c['quantity'])  # Chuyển quantity thành int
            total_quantity += c['quantity']
            total_amount += c['quantity'] * c['price']
    return {
        'total_quantity': total_quantity,
        'total_amount': total_amount
    }


def add_receipt(cart, delivery_method, payment_method, phone, email, delivery_address=None):
    if cart:
        receipt = Receipt(
            user=current_user,
            delivery_method=DeliveryMethod[delivery_method.upper()],
            payment_method=PaymentMethod[payment_method.upper()],
            delivery_address=delivery_address,
            phone=phone,
            email=email
        )
        db.session.add(receipt)
        for c in cart.values():
            book = Book.query.get(c['id'])
            if not book or book.stock < c['quantity']:
                raise Exception(f"Sản phẩm {c['name']} không đủ số lượng trong kho!")
            d = ReceiptDetail(receipt=receipt,
                              product_id=c['id'],
                              quantity=c['quantity'],
                              unit_price=c['price'])
            db.session.add(d)
        db.session.commit()

def stats_by_category(month, year):
    """Thống kê doanh thu theo thể loại sách trong tháng của năm"""
    return db.session.query(
        BookCategory.name,
        func.sum(ReceiptDetail.quantity * ReceiptDetail.unit_price),
        func.count(ReceiptDetail.product_id),
        func.round(
            (func.sum(ReceiptDetail.quantity * ReceiptDetail.unit_price) * 100.0 /
             func.sum(func.sum(ReceiptDetail.quantity * ReceiptDetail.unit_price)).over()),2
        )
    ).join(Book, ReceiptDetail.product_id == Book.id
    ).join(BookCategory, Book.category_id == BookCategory.id
    ).join(Receipt, ReceiptDetail.receipt_id == Receipt.id
    ).filter(
        func.extract('month', Receipt.created_date) == month,
        func.extract('year', Receipt.created_date) == year
    ).group_by(BookCategory.name).all()

def stats_book_sold(month, year):
    """Thống kê tần suất sách bán trong tháng của năm"""
    return db.session.query(
        Book.name,
        BookCategory.name,
        func.sum(ReceiptDetail.quantity),
        func.round((func.sum(ReceiptDetail.quantity) * 100.0 /
             func.sum(func.sum(ReceiptDetail.quantity)).over()),2
        )
    ).join(ReceiptDetail, ReceiptDetail.product_id == Book.id
    ).join(BookCategory, Book.category_id == BookCategory.id
    ).join(Receipt, ReceiptDetail.receipt_id == Receipt.id
    ).filter(
        func.extract('month', Receipt.created_date) == month,
        func.extract('year', Receipt.created_date) == year
    ).group_by(Book.name,BookCategory.name).all()


def export_stats_to_excel(revenue_stats, book_stats, month, year, total_revenue):
    wb = Workbook()

    # Tạo sheet báo cáo doanh thu
    revenue_sheet = wb.active
    revenue_sheet.title = "Báo cáo doanh thu"

    # Định dạng tiêu đề
    header_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Tiêu đề báo cáo doanh thu
    revenue_sheet['A1'] = f"BÁO CÁO DOANH THU THEO THÁNG {month}/{year}"
    revenue_sheet.merge_cells('A1:E1')
    revenue_sheet['A1'].font = header_font
    revenue_sheet['A1'].alignment = Alignment(horizontal='center')

    # Header cho bảng doanh thu
    headers = ['STT', 'Thể loại sách', 'Doanh thu (VNĐ)', 'Số lượt bán', 'Tỷ lệ (%)']
    for col, header in enumerate(headers, 1):
        cell = revenue_sheet.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = header_fill

    # Thêm dữ liệu doanh thu
    for row, stat in enumerate(revenue_stats, 4):
        revenue_sheet.cell(row=row, column=1).value = row - 3  # STT
        revenue_sheet.cell(row=row, column=2).value = stat[0]  # Thể loại
        revenue_sheet.cell(row=row, column=3).value = float(stat[1])  # Doanh thu
        revenue_sheet.cell(row=row, column=4).value = stat[2]  # Số lượt bán
        revenue_sheet.cell(row=row, column=5).value = float(stat[3])  # Tỷ lệ

    # Tổng doanh thu
    total_row = len(revenue_stats) + 4
    revenue_sheet.cell(row=total_row, column=2).value = "Tổng doanh thu:"
    revenue_sheet.cell(row=total_row, column=3).value = float(total_revenue)
    revenue_sheet.cell(row=total_row, column=2).font = Font(bold=True)
    revenue_sheet.cell(row=total_row, column=3).font = Font(bold=True)

    # Tạo sheet báo cáo tần suất
    freq_sheet = wb.create_sheet("Báo cáo tần suất")

    # Tiêu đề báo cáo tần suất
    freq_sheet['A1'] = f"BÁO CÁO TẦN SUẤT SÁCH BÁN THÁNG {month}/{year}"
    freq_sheet.merge_cells('A1:E1')
    freq_sheet['A1'].font = header_font
    freq_sheet['A1'].alignment = Alignment(horizontal='center')

    # Header cho bảng tần suất
    freq_headers = ['STT', 'Tên sách', 'Thể loại', 'Số lượng', 'Tỷ lệ (%)']
    for col, header in enumerate(freq_headers, 1):
        cell = freq_sheet.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = header_fill

    # Thêm dữ liệu tần suất
    for row, stat in enumerate(book_stats, 4):
        freq_sheet.cell(row=row, column=1).value = row - 3  # STT
        freq_sheet.cell(row=row, column=2).value = stat[0]  # Tên sách
        freq_sheet.cell(row=row, column=3).value = stat[1]  # Thể loại
        freq_sheet.cell(row=row, column=4).value = float(stat[2])  # Số lượng
        freq_sheet.cell(row=row, column=5).value = float(stat[3])  # Tỷ lệ

    # Điều chỉnh độ rộng cột cho cả hai sheet
    for sheet in [revenue_sheet, freq_sheet]:
        for i in range(1, sheet.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(i)

            for cell in sheet[column_letter]:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                except:
                    pass

            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width

    # Định dạng số cho các cột số liệu
    for sheet in [revenue_sheet, freq_sheet]:
        # Định dạng cột doanh thu với phân cách hàng nghìn
        for row in range(4, sheet.max_row + 1):
            if sheet == revenue_sheet:
                revenue_cell = sheet.cell(row=row, column=3)
                revenue_cell.number_format = '#,##0'

            # Định dạng cột tỷ lệ với 2 số thập phân
            ratio_cell = sheet.cell(row=row, column=5)
            ratio_cell.number_format = '0.00%'

    return wb